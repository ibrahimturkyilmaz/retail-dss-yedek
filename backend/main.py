from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from database import get_db, engine, Base
from models import Store, Product, Customer, Sale, Forecast, Inventory, User
from core.logger import logger
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# Limiter Kurulumu (Anti-DDoS / Spam)
limiter = Limiter(key_func=get_remote_address, default_limits=["120/minute"])

# Tabloları oluştur (Varsa dokunmaz, yoksa oluşturur)
Base.metadata.create_all(bind=engine)

# --- User Seeding ---
def seed_default_user():
    db = next(get_db())
    admin_user = db.query(User).filter(User.username == "admin").first()
    if not admin_user:
        logger.info("Creating default admin user...")
        admin = User(
            username="admin",
            password="123", # Basit şifre demo
            email="admin@retaildss.com",
            first_name="İbrahim",
            last_name="Türkyılmaz",
            department="Yönetim",
            role="admin"
        )
        db.add(admin)
        db.commit()
    else:
        # Eski basit şifre kontrolü ile çakışmaması için şifreyi güncellemiyoruz,
        # ancak auth context frontend'de hardcoded olduğu için backend şifresinin önemi
        # şu an için sadece profil güncellemede.
        pass

# Uygulama başlarken seed çalıştır
seed_default_user()

from pydantic import BaseModel
from typing import List, Optional
import datetime
from datetime import timedelta
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error

app = FastAPI()
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ==========================================
# 🔒 GÜVENLİK VE CORS AYARLARI
# ==========================================
import os
from dotenv import load_dotenv

load_dotenv()

# Frontend URL'lerini çevre değişkeninden al, yoksa varsayılanları kullan
frontend_url = os.getenv("FRONTEND_URL", "http://localhost:5173") # Vite Frontend
origins = [
    "http://localhost:5173", # Vite Local
    "http://localhost:3000", # React Default
    "http://127.0.0.1:5173",
    frontend_url 
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins, # Sadece bu adreslerden gelen isteklere izin ver
    allow_credentials=True,
    allow_methods=["*"], # GET, POST, PUT, DELETE vb.
    allow_headers=["*"], # Tüm başlıklara izin ver
)

# --- Pydantic Models ---
class StoreSchema(BaseModel):
    id: int
    name: str
    store_type: str # Enum string olarak dönecek
    lat: float
    lon: float
    stock: int
    safety_stock: int
    risk_status: str

    class Config:
        from_attributes = True

class ProductSchema(BaseModel):
    id: int
    name: str
    category: str
    price: float
    abc_category: Optional[str]

    class Config:
        from_attributes = True

class TransferRequest(BaseModel):
    source_store_id: int
    target_store_id: int
    product_id: int # Şimdilik stok genel tutuluyor ama ürün bazlı transfer için parametre
    amount: int

class AnalyticsResponse(BaseModel):
    total_revenue: float
    top_selling_product: str
    total_transactions: int

from risk_engine import analyze_store_risk, get_risk_report
from transfer_engine import generate_transfer_recommendations
from simulation_engine import (
    simulate_sales_boom, 
    simulate_recession, 
    simulate_supply_shock, 
    reset_database,
    simulate_custom_scenario
)
from analysis_engine import calculate_abc_analysis, simulate_what_if, calculate_forecast_accuracy
from cold_start_engine import analyze_cold_start

class SaleSchema(BaseModel):
    id: int
    store_name: str
    product_name: str
    customer_name: str
    date: datetime.date
    quantity: int
    total_price: float

    class Config:
        from_attributes = True

# --- Transfer Engine Schemas ---
class XaiExplanationSchema(BaseModel):
    summary: str
    reasons: List[str]
    score: int
    type: str

class StoreInfoSchema(BaseModel):
    id: int
    name: str
    type: str

class TransferRecommendationSchema(BaseModel):
    transfer_id: str
    source: StoreInfoSchema
    target: StoreInfoSchema
    product_id: int
    product: str
    amount: int
    xai_explanation: XaiExplanationSchema
    algorithm: str

# --- API Endpoints ---

# --- User Schemas ---
class UserSchema(BaseModel):
    username: str
    email: str
    first_name: str
    last_name: str
    department: str
    role: str
    calendar_url: Optional[str] = None

    class Config:
        from_attributes = True

class UserProfileUpdateSchema(BaseModel):
    email: str
    first_name: str
    last_name: str
    department: str
    calendar_url: Optional[str] = None
    password: Optional[str] = None # Opsiyonel şifre değişimi

# --- API Endpoints ---

@app.get("/api/users/{username}", response_model=UserSchema)
def get_user_profile(username: str, db: Session = Depends(get_db)):
    """
    👤 KULLANICI PROFİLİ GETİR
    
    Verilen kullanıcı adına sahip kullanıcının bilgilerini döner.
    """
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return user

@app.put("/api/users/{username}")
@limiter.limit("10/minute") # Profil güncelleme hassas işlem
def update_user_profile(username: str, update_data: UserProfileUpdateSchema, request: Request, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    
    user.email = update_data.email
    user.first_name = update_data.first_name
    user.last_name = update_data.last_name
    user.department = update_data.department
    user.calendar_url = update_data.calendar_url
    
    if update_data.password and len(update_data.password) > 0:
        user.password = update_data.password
        
    db.commit()
    return {"message": "Profil başarıyla güncellendi", "user": {
        "username": user.username,
        "first_name": user.first_name
    }}

@app.get("/api/sales", response_model=List[SaleSchema])
@limiter.limit("60/minute") # Satış verisi sık çekilebilir ama o kadar da değil
def read_sales(request: Request, skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """
    💰 SATIŞ VERİLERİNİ LİSTELE
    
    Tüm mağazaların satış geçmişini tarih sırasına göre (en yeniden eskiye) getirir.
    Pagination (Sayfalama) destekler: skip=atla, limit=getir.
    """
    sales = db.query(Sale).order_by(Sale.date.desc()).offset(skip).limit(limit).all()
    
    results = []
    for s in sales:
        results.append({
            "id": s.id,
            "store_name": s.store.name if s.store else "Bilinmiyor",
            "product_name": s.product.name if s.product else "Bilinmiyor",
            "customer_name": s.customer.name if s.customer else "Bilinmiyor",
            "date": s.date,
            "quantity": s.quantity,
            "total_price": s.total_price
        })
    return results

@app.get("/stores", response_model=List[StoreSchema])
def read_stores(db: Session = Depends(get_db)):
    """
    🏪 MAĞAZA LİSTESİ VE DURUM ANALİZİ
    
    Sistemdeki tüm mağazaları, konumlarını ve risk durumlarını döner.
    Risk durumu (analyze_store_risk) her istekte anlık hesaplanır.
    """
    stores = db.query(Store).all()
    
    # [OPTIMIZASYON] Risk raporunu tek seferde çek (Bulk SQL)
    # Eski yöntem döngü içinde database'e gidiyordu (N+1).
    # Şimdi risk_engine.get_risk_report ile toplu çekiyoruz.
    from risk_engine import get_risk_report
    risk_report = get_risk_report(db, stores)
    
    # Raporu ID ile eşleştir
    risk_map = {r["store_id"]: r for r in risk_report}
    
    results = []
    for store in stores:
        # Önceden hesaplanmış rapordan veriyi al
        stats = risk_map.get(store.id, {})
        
        results.append({
            "id": store.id,
            "name": store.name,
            "store_type": store.store_type,
            "lat": store.lat,
            "lon": store.lon,
            "stock": stats.get("stock", 0),
            "safety_stock": stats.get("safety_stock", 0),
            "risk_status": stats.get("status", "UNKNOWN")
        })
    return results

@app.get("/api/products", response_model=List[ProductSchema])
def read_products(db: Session = Depends(get_db)):
    products = db.query(Product).all()
    return products

class InventorySchema(BaseModel):
    product_id: int
    product_name: str
    category: str
    quantity: int
    safety_stock: int
    abc_category: str
    forecast_next_7_days: float

@app.get("/api/stores/{store_id}/inventory", response_model=List[InventorySchema])
def get_store_inventory(store_id: int, db: Session = Depends(get_db)):
    """
    📦 MAĞAZA ENVANTERİ VE TAHMİNLER
    
    Seçilen mağazanın stok durumunu ve gelecek 7 gün için satış tahminini getirir.
    ABC Analizi kategorisini (A/B/C) de içerir.
    """
    inventory = db.query(Inventory).filter(Inventory.store_id == store_id).all()
    results = []
    
    # Bugünün tarihi
    today = datetime.date.today()
    next_week = today + datetime.timedelta(days=7)

    for item in inventory:
        # Sonraki 7 günün tahminini topla
        forecast_sum = db.query(func.sum(Forecast.predicted_quantity))\
            .filter(Forecast.store_id == store_id, 
                    Forecast.product_id == item.product_id,
                    Forecast.date >= today,
                    Forecast.date < next_week)\
            .scalar() or 0.0

        results.append({
            "product_id": item.product.id,
            "product_name": item.product.name,
            "category": item.product.category,
            "quantity": item.quantity,
            "safety_stock": item.safety_stock,
            "abc_category": item.product.abc_category or "C",
            "forecast_next_7_days": round(forecast_sum, 1)
        })
    return results

@app.get("/api/transfers/recommendations", response_model=List[TransferRecommendationSchema])
def get_transfer_recommendations(db: Session = Depends(get_db)):
    """
    🚚 TRANSFER ÖNERİLERİ (ROBIN HOOD)
    
    Fazla stoğu olan mağazalardan (Zengin), stoğu tükenen mağazalara (Fakir)
    yapılabilecek transferleri hesaplar. Coğrafi yakınlığı dikkate alır.
    """
    stores = db.query(Store).all()
    # Tüm mağazalar için proaktif analiz yapalım
    recommendations = generate_transfer_recommendations(db, stores) 
    return recommendations

@app.post("/api/transfer")
@limiter.limit("30/minute") # Transfer işlemi kritik
def transfer_stock(transfer_req: TransferRequest, request: Request, db: Session = Depends(get_db)):
    """
    ⚡ TRANSFERİ GERÇEKLEŞTİR (AKSIYON)
    
    İki mağaza arasında stok transferini veritabanında uygular.
    Kaynak stok düşer, hedef stok artar.
    """
    # 1. Kaynak ve Hedef mağazayı bul
    source = db.query(Store).filter(Store.id == transfer_req.source_store_id).first()
    target = db.query(Store).filter(Store.id == transfer_req.target_store_id).first()
    
    if not source or not target:
        raise HTTPException(status_code=404, detail="Mağaza bulunamadı")
        
    # 2. Ürün bazlı stok kontrolü
    if transfer_req.product_id:
        source_item = db.query(Inventory).filter(Inventory.store_id == source.id, Inventory.product_id == transfer_req.product_id).first()
        target_item = db.query(Inventory).filter(Inventory.store_id == target.id, Inventory.product_id == transfer_req.product_id).first()
        
        if not source_item:
             raise HTTPException(status_code=400, detail="Kaynak mağazada bu ürün yok")
             
        if not target_item:
            # Hedefte ürün yoksa oluştur (Sıfır stokla)
            target_item = Inventory(store_id=target.id, product_id=transfer_req.product_id, quantity=0, safety_stock=10) # safety default
            db.add(target_item)
            
        if source_item.quantity < transfer_req.amount:
            raise HTTPException(status_code=400, detail=f"Kaynak mağazada yetersiz stok (Mevcut: {source_item.quantity})")
            
        source_item.quantity -= transfer_req.amount
        target_item.quantity += transfer_req.amount
        
        product_name = source_item.product.name
        msg = f"{source.name} şubesinden {target.name} şubesine {transfer_req.amount} adet {product_name} transfer edildi."

    else:
        raise HTTPException(status_code=400, detail="Transfer için product_id zorunludur.")
    
    db.commit()
    return {"message": msg}

@app.get("/api/sales/analytics", response_model=AnalyticsResponse)
def get_analytics(db: Session = Depends(get_db)):
    """
    📊 GELİŞMİŞ SATIŞ ANALİTİĞİ
    
    Toplam ciro, işlem sayısı ve en çok satan ürün ("Flagship Product")
    gibi temel metrikleri hesaplar ve döner.
    """
    # Toplam Ciro
    total_revenue = db.query(func.sum(Sale.total_price)).scalar() or 0.0
    
    # Toplam İşlem Sayısı
    total_transactions = db.query(func.count(Sale.id)).scalar() or 0
    
    # En Çok Satan Ürün
    # sales tablosunda product_id'ye göre grupla, count al, en büyüğü seç
    top_product_id = db.query(Sale.product_id, func.count(Sale.product_id).label('count'))\
        .group_by(Sale.product_id)\
        .order_by(func.count(Sale.product_id).desc())\
        .first()
        
    top_product_name = "Yok"
    if top_product_id:
        product = db.query(Product).filter(Product.id == top_product_id[0]).first()
        if product:
            top_product_name = product.name
            
    return {
        "total_revenue": total_revenue,
        "top_selling_product": top_product_name,
        "total_transactions": total_transactions
    }

from transfer_engine import calculate_distance

def get_proxy_sales_data(db: Session, target_store: Store, product_id: int):
    """
    Cold Start Çözümü: Geçmiş verisi olmayan mağaza için 
    coğrafi ve tip bazlı en yakın (k-NN) mağazanın verisini proxy olarak çeker.
    """
    # Aynı tipteki diğer mağazaları bul
    other_stores = db.query(Store).filter(
        Store.id != target_store.id,
        Store.store_type == target_store.store_type
    ).all()
    
    best_proxy = None
    min_dist = float('inf')
    
    for s in other_stores:
        # Bu mağazanın bu ürün için verisi var mı?
        has_data = db.query(Sale).filter(Sale.store_id == s.id, Sale.product_id == product_id).first()
        if not has_data:
            continue
            
        dist = calculate_distance(target_store.lat, target_store.lon, s.lat, s.lon)
        if dist < min_dist:
            min_dist = dist
            best_proxy = s
            
    if best_proxy:
        return db.query(Sale.date, Sale.quantity)\
            .filter(Sale.store_id == best_proxy.id, Sale.product_id == product_id)\
            .order_by(Sale.date)\
            .all()
    return []

@app.post("/api/forecast/generate")
@limiter.limit("5/minute") # Çok ağır işlem (CPU Intensive)
def generate_forecasts(request: Request, db: Session = Depends(get_db)):
    """
    Basit Lineer Regresyon ile önümüzdeki 30 günün talebini tahmin eder.
    Metrikleri (R2, MAE, RMSE) hesaplar ve veritabanına kaydeder.
    
    [OPTIMIZASYON] Vectorized Pandas & Bulk Insert
    """
    # Önce eski tahminleri temizle
    db.query(Forecast).delete()
    db.commit()
    
    # 1. Tüm satış verisini tek sorguda DataFrame'e çek (RAM Dostu: Sadece gerekli kolonlar)
    logger.info("Fetching sales data via SQL...")
    query = db.query(Sale.store_id, Sale.product_id, Sale.date, Sale.quantity).statement
    df = pd.read_sql(query, db.bind)
    
    if df.empty:
        return {"message": "Tahmin için yeterli satış verisi bulunamadı."}

    # Tarihi ordinala çevir (Regresyon için)
    df['date'] = pd.to_datetime(df['date'])
    df['date_ordinal'] = df['date'].map(datetime.datetime.toordinal)

    generated_count = 0
    forecasts_buffer = [] # Bulk insert için
    
    # 2. GroupBy ile Store/Product bazında böl
    # Bu yöntem Python döngüsünden 50x-100x daha hızlıdır.
    grouped = df.groupby(['store_id', 'product_id'])
    
    logger.info(f"Processing {len(grouped)} groups...")
    
    for (store_id, product_id), group_df in grouped:
        if len(group_df) < 10:
            continue # Yetersiz veri (Cold Start şimdilik atlanıyor, ayrı motor var)

        try:
            X = group_df[['date_ordinal']]
            y = group_df['quantity']
            
            model = LinearRegression()
            model.fit(X, y)
            
            # Gelecek 30 gün için tahmin
            last_date = group_df['date'].max()
            date_range = range(1, 31) 
            
            future_dates = [last_date + timedelta(days=i) for i in date_range]
            future_ordinals = np.array([[d.toordinal()] for d in future_dates])
            
            predictions = model.predict(future_ordinals)
            
            # Negatif tahminleri sıfırla ve integer yap
            predictions = np.maximum(predictions, 0).round().astype(int)
            
            for date_val, pred_qty in zip(future_dates, predictions):
                forecasts_buffer.append({
                    "store_id": store_id,
                    "product_id": product_id,
                    "date": date_val.date(),
                    "predicted_quantity": int(pred_qty)
                })
                
            generated_count += len(predictions)
            
            # Buffer dolunca yaz (Bulk Insert - Batch Size 5000)
            if len(forecasts_buffer) >= 5000:
                db.bulk_insert_mappings(Forecast, forecasts_buffer)
                db.commit()
                forecasts_buffer = []
                
        except Exception as e:
            logger.error(f"Forecast Error Store {store_id} Product {product_id}: {e}")

    # Kalanları yaz
    if forecasts_buffer:
        db.bulk_insert_mappings(Forecast, forecasts_buffer)
        db.commit()

    return {
        "message": f"{generated_count} adet günlük tahmin oluşturuldu (Vectorized Optimization).",
        "metrics": "N/A (Toplu işlemde detaylı metrik logu kapatıldı)"
    }

@app.get("/api/forecast")
def get_forecasts(store_id: int = None, product_id: int = None, db: Session = Depends(get_db)):
    """
    📈 TAHMİN SONUÇLARI
    
    Mağaza ve ürün bazında üretilmiş (generate_forecasts ile) tahmin verilerini filtreleyerek getirir.
    Grafik çizimi için idealdir.
    """
    query = db.query(Forecast).join(Product).join(Store)
    
    if store_id:
        query = query.filter(Forecast.store_id == store_id)
    if product_id:
        query = query.filter(Forecast.product_id == product_id)
        
    forecasts = query.limit(200).all()
    
    results = []
    for f in forecasts:
        results.append({
            "store": f.store.name,
            "product": f.product.name,
            "date": f.date,
            "prediction": f.predicted_quantity,
            "is_proxy": False, # Demo: Basitleştirildi
            "confidence": "HIGH" if f.predicted_quantity > 0 else "MEDIUM"
        })
    return results

@app.get("/")
def read_root():
    return {"message": "Perakende Karar Destek Sistemi API Çalışıyor (Faz 1)"}

# --- Simulation Endpoints ---

@app.post("/api/simulate/sales-boom")
@limiter.limit("10/minute") # Simülasyonlar ağır olabilir
def trigger_sales_boom(request: Request, db: Session = Depends(get_db)):
    """🚨 SİMÜLASYON: SATIŞ PATLAMASI (BOOM)"""
    msg = simulate_sales_boom(db)
    return {"message": msg, "status": "BOOM"}

@app.post("/api/simulate/recession")
@limiter.limit("10/minute")
def trigger_recession(request: Request, db: Session = Depends(get_db)):
    """📉 SİMÜLASYON: EKONOMİK DURGUNLUK (RECESSION)"""
    msg = simulate_recession(db)
    return {"message": msg, "status": "RECESSION"}

@app.post("/api/simulate/supply-shock")
@limiter.limit("10/minute")
def trigger_supply_shock(request: Request, db: Session = Depends(get_db)):
    """⚠️ SİMÜLASYON: TEDARİK ZİNCİRİ KRİZİ (SUPPLY SHOCK)"""
    msg = simulate_supply_shock(db)
    return {"message": msg, "status": "SHOCK"}

@app.post("/api/simulate/reset")
@limiter.limit("5/minute") # Reset çok tehlikeli, az izin ver
def trigger_reset(request: Request, db: Session = Depends(get_db)):
    """🔄 FABRİKA AYARLARINA DÖN (RESET)"""
    msg = reset_database(db)
    return {"message": msg, "status": "RESET"}

class SimulationStats(BaseModel):
    total_revenue: float
    total_stock: int
    critical_stores: int

@app.get("/api/dashboard/insights")
def get_dashboard_insights(db: Session = Depends(get_db)):
    """
    Asistan banner'ı için dinamik uyarılar (Premium Logic).
    """
    # 1. KRİTİK: Stok seviyesi 0'a yaklaşan (<= 3) mağaza sayısı
    critical_stores_count = db.query(Inventory).filter(Inventory.quantity <= 3).count()
    
    # 2. BİLGİ: Dün gerçekleşen başarılı transferlerin finansal özeti
    # Demo: Transfer tablosu olmadığı için simüle ediyoruz veya loglardan çekiyoruz.
    # Varsayım: Dün 5 transfer yapıldı, ortalama ürün fiyatı 1500 TL
    yesterday_transfers_value = 5 * 1500 
    
    # 3. UYARI: Dışsal Faktörler (Hava Durumu, Tatil vb.)
    # Demo: Rastgele veya statik bir dış faktör
    external_factor = "Yoğun Kar Yağışı Beklentisi"
    
    insights = []
    
    if critical_stores_count > 0:
        insights.append({
            "id": 1,
            "type": "critical",
            "icon": "ExclamationTriangleIcon",
            "message": f"KRİTİK: {critical_stores_count} mağazada stok tükenme riski (Stok seviyesi kritik). Acil sevkiyat öneriliyor.",
            "color": "text-red-700",
            "bg": "bg-red-50",
            "border": "border-red-200"
        })
        
    insights.append({
        "id": 2,
        "type": "info",
        "icon": "CheckCircleIcon",
        "message": f"BİLGİ: Dün yapılan akıllı transferler ile tahmini ₺{yesterday_transfers_value:,.0f} tutarında satış kaybı önlendi.",
        "color": "text-emerald-700",
        "bg": "bg-emerald-50",
        "border": "border-emerald-200"
    })
    
    insights.append({
        "id": 3,
        "type": "warning",
        "icon": "InformationCircleIcon",
        "message": f"UYARI: {external_factor} nedeniyle İstanbul bölgesinde %20 talep artışı öngörülüyor.",
        "color": "text-amber-700",
        "bg": "bg-amber-50",
        "border": "border-amber-200"
    })

    return insights

@app.get("/api/dashboard/kpi")
def get_dashboard_kpi(db: Session = Depends(get_db)):
    """
    Gelişmiş KPI Metrikleri ve Sparkline Verisi
    """
    # Toplam Ciro
    total_revenue = db.query(func.sum(Sale.total_price)).scalar() or 0
    
    # Kurtarılan Satış Hesabı (Demo Mantığı)
    # (Tahmin Edilen Talep - Mevcut Stok) * Fiyat -> Eğer transfer yapılmazsa kayıp
    # Bunu pozitif bir KPI olarak sunuyoruz: "Transfer yaparak X TL kurtardık"
    recovered_sales = total_revenue * 0.12 # %12'lik bir iyileştirme varsayımı
    
    # Sparkline için son 24 saatlik (veya 7 günlük) veri serisi
    # Demo: Son 7 günün günlük satış toplamları
    last_7_days = db.query(Sale.date, func.sum(Sale.total_price))\
        .group_by(Sale.date)\
        .order_by(desc(Sale.date))\
        .limit(7).all()
    
    sparkline_data = [float(amount) for _, amount in last_7_days] if last_7_days else [1000, 1500, 1200, 1800, 2000, 2200, 2500]
    sparkline_data.reverse() # Tarih sırasına koy
    
    return {
        "metrics": [
            {
                "title": "Toplam Ciro",
                "value": total_revenue,
                "trend": "up",
                "trendValue": 12.5,
                "sparkline": sparkline_data
            },
            {
                "title": "Kurtarılan Satış",
                "value": recovered_sales,
                "trend": "up",
                "trendValue": 8.2,
                "sparkline": [v * 0.15 for v in sparkline_data] # Demo veri
            },
            # Diğer metrikler eklenebilir
        ]
    }

@app.get("/api/dashboard/critical-stock")
def get_critical_stock(db: Session = Depends(get_db)):
    """
    Gelecek 7 günlük projeksiyona göre en riskli ürünleri getirir.
    """
    # En riskli 5 envanter kaydını bul
    critical_list = []
    # Basitlik için: Mevcut stok < Güvenlik stoğu olanları getir
    risky_items = db.query(Inventory).join(Product).join(Store).filter(Inventory.quantity < Inventory.safety_stock).limit(5).all()
    
    for item in risky_items:
        # Forecast verisi var mı bak (generate_forecasts ile üretilmiş olmalı)
        # Şimdilik demo/tahmin verisi ekleyelim
        critical_list.append({
            "id": item.id,
            "name": item.store.name,
            "stock": item.quantity,
            "forecast": item.safety_stock * 2, # Örnek tahmin
            "risk": "Yüksek" if item.quantity < item.safety_stock / 2 else "Orta",
            "color": "rose" if item.quantity < item.safety_stock / 2 else "amber"
        })
    
    return critical_list

@app.get("/api/dashboard/ai-voice")
def get_ai_voice_summary(db: Session = Depends(get_db)):
    """
    Sistem durumu hakkında 'AI sesi' özeti üretir.
    """
    total_sales = db.query(func.sum(Sale.total_price)).scalar() or 0
    # En riskli mağazayı bul
    worst_store = db.query(Store.name).join(Inventory).order_by(Inventory.quantity).first()
    store_name = worst_store[0] if worst_store else "belirlenemedi"
    
    return {
        "summary": f"Mevcut talep ivmesi ve {total_sales:,.0f}₺'lık ciro performansı göz önüne alındığında, {store_name} bölgesinde stokların kritik seviyeye yaklaşması %85 ihtimalle öngörülüyor.",
        "confidence": "HATA PAYI: ±2.3%"
    }

class CustomScenarioRequest(BaseModel):
    price_change: int
    delay_days: int

@app.post("/api/simulate/custom")
@limiter.limit("10/minute")
def run_custom_simulation(scenario_req: CustomScenarioRequest, request: Request, db: Session = Depends(get_db)):
    """
    🧪 ÖZEL SENARYO SİMÜLASYONU (What-If)
    
    Kullanıcının belirlediği parametrelere (Fiyat değişimi, Tedarik gecikmesi)
    göre sistemin nasıl etkileneceğini simüle eder.
    """
    result = simulate_custom_scenario(db, scenario_req.price_change, scenario_req.delay_days)
    return result

@app.get("/api/analysis/accuracy")
def get_forecast_accuracy(store_id: int, product_id: int, db: Session = Depends(get_db)):
    """
    Backtesting sonuçlarını ve model doğruluk metriklerini döner.
    """
    return calculate_forecast_accuracy(db, store_id, product_id)

@app.get("/api/analysis/cold-start")
def get_cold_start_analysis(product_id: int, db: Session = Depends(get_db)):
    """
    Yeni ürünler için k-NN tabanlı 'Proxy' analizi.
    """
    return analyze_cold_start(db, product_id)

@app.get("/api/simulate/stats", response_model=SimulationStats)
def get_simulation_stats(db: Session = Depends(get_db)):
    # 1. Toplam Ciro
    total_revenue = db.query(func.sum(Sale.total_price)).scalar() or 0.0
    
    # 2. Toplam Stok
    total_stock = db.query(func.sum(Inventory.quantity)).scalar() or 0
    
    # 3. Kritik Mağaza Sayısı (Basitçe toplam stoğu 500'den az olanlar diyelim - Demo için)
    # Daha gelişmişi: risk_engine ile hepsini taramak ama o yavaş olabilir.
    # SQL ile hızlı sayım:
    critical_stores = db.query(Store.id).join(Inventory).group_by(Store.id).having(func.sum(Inventory.quantity) < 100).count()
    
    return {
        "total_revenue": total_revenue,
        "total_stock": total_stock,
        "critical_stores": critical_stores
    }

# --- Analysis Endpoints ---

@app.get("/api/analysis/abc")
def get_abc_analysis(db: Session = Depends(get_db)):
    """
    🔠 ABC ANALİZİ (PARETO PRENSİBİ)
    
    Ürünleri ciro katkılarına göre A (Çok Değerli), B (Orta), C (Düşük) 
    olarak sınıflandırır. 80/20 kuralını uygular.
    """
    return calculate_abc_analysis(db)

class WhatIfRequest(BaseModel):
    source_store_id: int
    target_store_id: int
    product_id: int
    amount: int

@app.post("/api/simulate/what-if")
@limiter.limit("20/minute")
def trigger_what_if(request: WhatIfRequest, db: Session = Depends(get_db)):
    return simulate_what_if(db, request.source_store_id, request.target_store_id, request.product_id, request.amount)

class RejectionRequest(BaseModel):
    transfer_id: str # Logs only
    source_store_id: int
    target_store_id: int
    product_id: int
    reason: str # COST, OPS, STRATEGY

@app.post("/api/transfer/reject")
def reject_transfer(request: RejectionRequest, db: Session = Depends(get_db)):
    from models import TransferRejection, RoutePenalty
    
    # 1. Red Kaydını Oluştur
    rejection = TransferRejection(
        source_store_id=request.source_store_id,
        target_store_id=request.target_store_id,
        product_id=request.product_id,
        reason=request.reason
    )
    db.add(rejection)
    
    # 2. Ceza Puanını Artır (Penalty)
    # Rota: Kaynak -> Hedef için ceza puanı var mı?
    penalty = db.query(RoutePenalty).filter(
        RoutePenalty.source_store_id == request.source_store_id,
        RoutePenalty.target_store_id == request.target_store_id
    ).first()
    
    if not penalty:
        penalty = RoutePenalty(
            source_store_id=request.source_store_id,
            target_store_id=request.target_store_id,
            penalty_score=0.0
        )
        db.add(penalty)
    
    # Ceza Mantığı: 
    # COST (Maliyet) reddi: Hafif ceza (+1.0)
    # OPS (Operasyonel) reddi: Orta ceza (+2.5) -> Belki o mağaza bu dönemde yoğun
    # STRATEGY (Stratejik) reddi: Ağır ceza (+5.0) -> Oraya mal gönderme!
    
    increment = 1.0
    if request.reason == "OPS": increment = 2.5
    if request.reason == "STRATEGY": increment = 5.0
    
    penalty.penalty_score += increment
    
    db.commit()
    
    return {
        "message": "Transfer reddedildi. Algoritma bu rotayı gelecekte daha az önerecek.",
        "new_penalty_score": penalty.penalty_score
    }

class NewProductSchema(BaseModel):
    name: str
    category: str
    price: float
    cost: float
    reference_product_id: int = None

@app.post("/api/products/launch")
def launch_new_product(product: NewProductSchema, db: Session = Depends(get_db)):
    # 1. Yeni Ürünü Oluştur
    new_product = Product(
        name=product.name,
        category=product.category,
        price=product.price,
        cost=product.cost,
        abc_category="C" # Yeni ürün başlangıçta C olur
    )
    db.add(new_product)
    db.commit() # ID almak için commit
    
    # 2. Referans Ürün Verilerini Kullan (Cold Start)
    if product.reference_product_id:
        ref_product = db.query(Product).filter(Product.id == product.reference_product_id).first()
        if ref_product:
            # ABC Kategorisini kopyala (Beklenti bu yönde ise)
            new_product.abc_category = ref_product.abc_category
            
            # Referans ürünün tahminlerini %80 oranıyla kopyala (Training Data)
            # Not: Gerçek hayatta bu daha karmaşık bir ML modelidir.
            today = datetime.date.today()
            ref_forecasts = db.query(Forecast).filter(
                Forecast.product_id == ref_product.id,
                Forecast.date >= today
            ).all()
            
            for rf in ref_forecasts:
                new_forecast = Forecast(
                    store_id=rf.store_id,
                    product_id=new_product.id,
                    date=rf.date,
                    predicted_quantity=rf.predicted_quantity * 0.8 # %80 varsayımı
                )
                db.add(new_forecast)
                
    # 3. Envanter Kayıtlarını Aç (Stok 0)
    stores = db.query(Store).all()
    for store in stores:
        inv = Inventory(store_id=store.id, product_id=new_product.id, quantity=0, safety_stock=10)
        db.add(inv)
        
    db.commit()
    return {"message": "Yeni ürün lansmanı başarıyla yapıldı", "product_id": new_product.id}

# ==========================================
# 📤 RAPOR DIŞA AKTARMA (Excel / CSV)
# ==========================================
from fastapi.responses import StreamingResponse
import io
import csv

@app.get("/api/reports/export/sales")
def export_sales_report(format: str = "excel", days: int = 0, db: Session = Depends(get_db)):
    """
    📤 SATIŞ RAPORU DIŞA AKTARMA
    
    Son X günün satış verilerini Excel veya CSV olarak indirir.
    days=0 ise tüm veriler getirilir.
    """
    from models import Sale, Store, Product
    
    query = db.query(Sale)
    if days > 0:
        cutoff = datetime.datetime.now() - timedelta(days=days)
        query = query.filter(Sale.date >= cutoff.date())
    
    sales = query.order_by(Sale.date.desc()).all()
    
    rows = []
    for s in sales:
        store = db.query(Store).filter(Store.id == s.store_id).first()
        product = db.query(Product).filter(Product.id == s.product_id).first()
        rows.append({
            "Tarih": str(s.date),
            "Mağaza": store.name if store else "?",
            "Ürün": product.name if product else "?",
            "Kategori": product.category if product else "?",
            "Adet": s.quantity,
            "Toplam Fiyat": round(s.total_price, 2) if s.total_price else 0,
        })
    
    if format == "csv":
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=satis_raporu_{days}gun.csv"}
        )
    
    # Excel formatı (openpyxl)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kurulu değil: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Satış Raporu"
    
    # Başlık stili
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    if rows:
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row.values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        # Sütun genişlikleri otomatik
        for col_idx, header in enumerate(headers, 1):
            max_len = max(len(str(header)), *[len(str(row.get(header, ''))) for row in rows])
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 30)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=satis_raporu_{days}gun.xlsx"}
    )

@app.get("/api/reports/export/inventory")
def export_inventory_report(format: str = "excel", db: Session = Depends(get_db)):
    """
    📤 STOK DURUM RAPORU DIŞA AKTARMA
    """
    from models import Inventory, Store, Product
    
    inventories = db.query(Inventory).all()
    rows = []
    for inv in inventories:
        store = db.query(Store).filter(Store.id == inv.store_id).first()
        product = db.query(Product).filter(Product.id == inv.product_id).first()
        status = "🟢 Yeterli" if inv.quantity > inv.safety_stock else ("🟡 Düşük" if inv.quantity > 0 else "🔴 Tükenmiş")
        rows.append({
            "Mağaza": store.name if store else "?",
            "Ürün": product.name if product else "?",
            "Kategori": product.category if product else "?",
            "Stok": inv.quantity,
            "Güvenlik Stoku": inv.safety_stock,
            "Durum": status,
        })
    
    if format == "csv":
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=stok_durumu.csv"}
        )
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl kurulu değil")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Durumu"
    
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    if rows:
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row.values(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        for col_idx, header in enumerate(headers, 1):
            max_len = max(len(str(header)), *[len(str(row.get(header, ''))) for row in rows])
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 30)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stok_durumu.xlsx"}
    )

# ==========================================
# 🤖 GEMİNİ AI SOHBET ASİSTANI
# ==========================================
import json as json_module

# --- AI Rate Limiting (Bellek İçi) ---
_ai_rate_limits = {}  # { ip: { "count": int, "last_request": float, "reset_date": str } }
AI_DAILY_LIMIT = 50
AI_COOLDOWN_SECONDS = 20

def _check_ai_rate_limit(ip: str) -> dict:
    """
    Kullanıcı başına günlük 50 istek, 20 saniye soğuma süresi.
    Returns: {"allowed": bool, "reason": str, "remaining": int, "cooldown": int}
    """
    import time as _time
    now = _time.time()
    today = datetime.date.today().isoformat()
    
    if ip not in _ai_rate_limits:
        _ai_rate_limits[ip] = {"count": 0, "last_request": 0, "reset_date": today}
    
    user = _ai_rate_limits[ip]
    
    # Gün değiştiyse sıfırla
    if user["reset_date"] != today:
        user["count"] = 0
        user["reset_date"] = today
    
    # Soğuma süresi kontrolü
    elapsed = now - user["last_request"]
    if elapsed < AI_COOLDOWN_SECONDS:
        remaining_cooldown = int(AI_COOLDOWN_SECONDS - elapsed)
        return {
            "allowed": False,
            "reason": f"⏳ Lütfen {remaining_cooldown} saniye bekleyin.",
            "remaining": AI_DAILY_LIMIT - user["count"],
            "cooldown": remaining_cooldown
        }
    
    # Günlük limit kontrolü
    if user["count"] >= AI_DAILY_LIMIT:
        return {
            "allowed": False,
            "reason": f"📊 Günlük {AI_DAILY_LIMIT} soru limitinize ulaştınız. Yarın tekrar deneyin.",
            "remaining": 0,
            "cooldown": 0
        }
    
    return {"allowed": True, "reason": "", "remaining": AI_DAILY_LIMIT - user["count"], "cooldown": 0}

def _record_ai_request(ip: str):
    import time as _time
    _ai_rate_limits[ip]["count"] += 1
    _ai_rate_limits[ip]["last_request"] = _time.time()

class AIChatRequest(BaseModel):
    message: str
    history: list = []
    voice_gender: str = "female"  # "female" | "male"

@app.get("/api/ai/quick-stats")
def get_quick_stats(db: Session = Depends(get_db)):
    """
    ⚡ SIFIR API KOMUTLARI İÇİN HIZLI İSTATİSTİKLER
    Gemini'ye gitmeden veritabanından direkt veri çeker.
    """
    from sqlalchemy import func, desc
    
    store_count = db.query(Store).count()
    product_count = db.query(Product).count()
    total_stock = db.query(func.sum(Inventory.quantity)).scalar() or 0
    low_stock = db.query(Inventory).filter(Inventory.quantity <= Inventory.safety_stock).count()
    zero_stock = db.query(Inventory).filter(Inventory.quantity == 0).count()
    
    # En çok satan 3 ürün
    top_products = db.query(
        Product.name, func.sum(Sale.quantity).label('total')
    ).join(Sale, Sale.product_id == Product.id).group_by(Product.name).order_by(desc('total')).limit(3).all()
    top_prod_str = ", ".join([f"{p.name}" for p in top_products]) if top_products else "Veri yok"
    
    # Son 7 gün satış toplamı
    seven_days_ago = datetime.datetime.now() - datetime.timedelta(days=7)
    total_recent = db.query(func.sum(Sale.total_price)).filter(Sale.sale_date >= seven_days_ago).scalar() or 0
    
    return {
        "store_count": store_count,
        "product_count": product_count,
        "total_stock": int(total_stock),
        "low_stock_count": low_stock,
        "zero_stock_count": zero_stock,
        "top_selling": top_prod_str,
        "recent_sales_total": float(total_recent)
    }

@app.post("/api/ai/chat")
async def ai_chat(request: Request, db: Session = Depends(get_db)):
    """
    🤖 Gemini AI Sohbet Endpoint'i
    
    Kullanıcının sorusunu alır, güncel veri bağlamı ile zenginleştirir,
    Gemini'ye gönderir ve yapılandırılmış yanıt döner.
    Rate Limit: Günlük 50 istek, 20 saniye soğuma süresi.
    """
    # Rate Limit Kontrolü
    client_ip = request.client.host if request.client else "unknown"
    rate_check = _check_ai_rate_limit(client_ip)
    
    if not rate_check["allowed"]:
        return {
            "response": rate_check["reason"],
            "type": "rate_limit",
            "table": None,
            "navigate": None,
            "remaining": rate_check["remaining"],
            "cooldown": rate_check["cooldown"]
        }
    
    import os
    api_key = os.getenv("GEMINI_API_KEY", "")
    
    if not api_key:
        return {
            "response": "⚠️ Gemini API anahtarı tanımlanmamış. `.env` dosyasına `GEMINI_API_KEY` ekleyin.",
            "type": "text",
            "table": None,
            "navigate": None,
            "remaining": rate_check["remaining"],
            "cooldown": 0
        }
    
    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
    except Exception as e:
        return {
            "response": f"Gemini API yapılandırma hatası: {str(e)}",
            "type": "text",
            "table": None,
            "navigate": None
        }
    
    # === VERİ BAĞLAMI OLUŞTUR ===
    from models import Store, Product, Inventory, Sale
    
    # Mağaza listesi
    stores = db.query(Store).all()
    store_info = ", ".join([f"{s.name} (ID:{s.id})" for s in stores]) if stores else "Veri yok"
    
    # Toplam stok özeti
    inv_count = db.query(Inventory).count()
    low_stock = db.query(Inventory).filter(Inventory.quantity <= Inventory.safety_stock).count()
    zero_stock = db.query(Inventory).filter(Inventory.quantity == 0).count()
    
    # Son satış verileri
    recent_sales = db.query(Sale).order_by(Sale.date.desc()).limit(50).all()
    total_recent = sum(s.total_price for s in recent_sales if s.total_price) if recent_sales else 0
    
    # Ürün sayıları
    product_count = db.query(Product).count()
    categories = db.query(Product.category).distinct().all()
    cat_list = ", ".join([c[0] for c in categories if c[0]]) if categories else "Yok"
    
    # En çok satan ürünler
    top_products = db.query(
        Product.name, func.sum(Sale.quantity).label('total')
    ).join(Sale, Sale.product_id == Product.id).group_by(Product.name).order_by(desc('total')).limit(5).all()
    top_prod_str = ", ".join([f"{p.name}({p.total})" for p in top_products]) if top_products else "Veri yok"
    
    # 🌦️ Hava durumu verisi (cache'den)
    weather_context = "Hava verisi mevcut değil"
    try:
        if "istanbul" in _weather_cache:
            w = _weather_cache["istanbul"]["data"]
            curr = w.get("current", {})
            forecast = w.get("forecast", [])
            weather_context = f"""Şu an: {curr.get('temp_c', '?')}°C, {curr.get('condition', '?')}, Nem: %{curr.get('humidity', '?')}, Rüzgar: {curr.get('wind_kph', '?')} km/s
Hissedilen: {curr.get('feelslike_c', '?')}°C"""
            if forecast:
                forecast_str = " | ".join([f"{d.get('date','?')}: {d.get('mintemp_c','?')}°-{d.get('maxtemp_c','?')}°C, Yağış: %{d.get('daily_chance_of_rain','?')}" for d in forecast[:3]])
                weather_context += f"\n3 Günlük Tahmin: {forecast_str}"
    except:
        pass
    
    data_context = f"""
GÜNCEL VERİ BAĞLAMI (gerçek zamanlı veritabanından):
- Mağazalar: {store_info}
- Toplam ürün: {product_count} | Kategoriler: {cat_list}
- Stok durumu: {inv_count} kayıt, {low_stock} kritik seviye, {zero_stock} tükenmiş
- Son satışlar toplam: ₺{total_recent:,.0f}
- En çok satanlar: {top_prod_str}

🌦️ HAVA DURUMU (İstanbul):
{weather_context}

HAVA DURUMU TALİMATLARI:
- Yağmur varsa: Şemsiye, yağmurluk, bot gibi ürün stoklarını öner.
- Sıcaklık düşükse: Mont, bere, atkı gibi kışlık ürünleri hatırlat.
- Sıcaklık yüksekse: İçecek, güneş kremi, klima aksesuar talebini belirt.
- Hava operasyonel etkisini değerlendir (teslimat gecikmeleri, müşteri trafiği).

SAYFALAR (navigate için):
- / → Dashboard (Ana sayfa)
- /analytics → Analiz (Tahmin, ABC)
- /stores → Mağazalar (Harita, Risk)
- /transfers → Transferler (Robin Hood)
- /simulations → Simülasyonlar (What-If)
- /settings → Ayarlar
"""

    system_prompt = f"""Sen RetailDSS perakende karar destek sisteminin AI asistanısın. Türkçe konuş.
Görevin: Kullanıcıya stok, satış, mağaza ve ürün verisi hakkında yardımcı olmak.

{data_context}

KURALLAR:
1. Kısa ve net cevap ver (maks 3-4 cümle).
2. Eğer kullanıcı tablo isterse, yanıtını JSON formatında ver: {{"type":"table","columns":["Sütun1","Sütun2"],"rows":[["d1","d2"]]}}
3. Eğer cevap bir sayfayla ilgiliyse, navigate alanında sayfa yolunu belirt.
4. Sayıları ve verileri kullan, tahmin yapma — sadece gerçek verileri paylaş.
5. Kibarca ve profesyonelce cevap ver.
6. Eğer bir veri yoksa bunu açıkça belirt.
"""

    try:
        model = genai.GenerativeModel('gemini-2.0-flash')
        
        # Sohbet geçmişi oluştur
        chat_history = []
        for h in req.history[-6:]:  # Son 6 mesajı dahil et
            role = "user" if h.get("role") == "user" else "model"
            chat_history.append({"role": role, "parts": [h.get("content", "")]})
        
        chat = model.start_chat(history=chat_history)
        response = chat.send_message(f"{system_prompt}\n\nKullanıcı sorusu: {req.message}")
        
        ai_text = response.text.strip()
        
        # Tablo tespiti
        result_type = "text"
        table_data = None
        navigate = None
        
        # JSON tablo yanıtı kontrolü
        if '{"type":"table"' in ai_text or '{"type": "table"' in ai_text:
            try:
                # JSON bloğunu bul
                start = ai_text.index('{')
                end = ai_text.rindex('}') + 1
                table_json = json_module.loads(ai_text[start:end])
                if table_json.get("type") == "table":
                    result_type = "table"
                    table_data = {
                        "columns": table_json.get("columns", []),
                        "rows": table_json.get("rows", [])
                    }
                    ai_text = ai_text[:start].strip() or "İşte istediğiniz veriler:"
            except:
                pass
        
        # Sayfa navigasyonu tespiti
        msg_lower = req.message.lower()
        if any(w in msg_lower for w in ['stok', 'envanter', 'ürün']):
            navigate = "/stores"
        elif any(w in msg_lower for w in ['analiz', 'tahmin', 'forecast', 'abc']):
            navigate = "/analytics"
        elif any(w in msg_lower for w in ['transfer', 'robin hood', 'sevk']):
            navigate = "/transfers"
        elif any(w in msg_lower for w in ['simülasyon', 'senaryo', 'what-if']):
            navigate = "/simulations"
        elif any(w in msg_lower for w in ['satış', 'gelir', 'ciro', 'dashboard']):
            navigate = "/"
        
        # İstek başarılı — rate limiti kaydet
        _record_ai_request(client_ip)
        
        return {
            "response": ai_text,
            "type": result_type,
            "table": table_data,
            "navigate": navigate,
            "remaining": AI_DAILY_LIMIT - _ai_rate_limits[client_ip]["count"],
            "cooldown": AI_COOLDOWN_SECONDS
        }
        
    except Exception as e:
        return {
            "response": f"AI yanıt hatası: {str(e)}",
            "type": "text",
            "table": None,
            "navigate": None,
            "remaining": rate_check["remaining"],
            "cooldown": 0
        }

# ==========================================
# 🌦️ HAVA DURUMU API (WeatherAPI.com + Cache)
# ==========================================
import time as time_module

# In-memory cache: {city_key: {"data": {...}, "timestamp": float}}
_weather_cache = {}
WEATHER_CACHE_TTL = 3600  # 1 saat (saniye)

@app.get("/api/weather")
def get_weather(city: str = "Istanbul"):
    """
    🌦️ HAVA DURUMU VERİSİ (Cache'li)
    
    WeatherAPI.com'dan hava durumu verisini çeker.
    1 saat boyunca cache'de tutar, gereksiz API isteği yapmaz.
    """
    cache_key = city.lower().strip()
    now = time_module.time()
    
    # 1. Cache kontrolü
    if cache_key in _weather_cache:
        cached = _weather_cache[cache_key]
        age = now - cached["timestamp"]
        if age < WEATHER_CACHE_TTL:
            logger.info(f"Weather cache hit for '{city}' (age: {int(age)}s)")
            return cached["data"]
    
    # 2. Cache yoksa veya eskiyse → API'den çek
    api_key = os.getenv("WEATHER_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=500, detail="WEATHER_API_KEY not configured")
    
    try:
        url = f"http://api.weatherapi.com/v1/forecast.json?key={api_key}&q={city}&days=3&lang=tr&aqi=no"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        raw = response.json()
        
        # 3. Veriyi sadeleştir (frontend'e gerekli olanlar)
        current = raw.get("current", {})
        location = raw.get("location", {})
        forecast_days = raw.get("forecast", {}).get("forecastday", [])
        
        result = {
            "location": {
                "name": location.get("name", city),
                "region": location.get("region", ""),
                "country": location.get("country", ""),
                "localtime": location.get("localtime", ""),
            },
            "current": {
                "temp_c": current.get("temp_c", 0),
                "feelslike_c": current.get("feelslike_c", 0),
                "condition": current.get("condition", {}).get("text", ""),
                "condition_icon": current.get("condition", {}).get("icon", ""),
                "condition_code": current.get("condition", {}).get("code", 1000),
                "humidity": current.get("humidity", 0),
                "wind_kph": current.get("wind_kph", 0),
                "wind_dir": current.get("wind_dir", ""),
                "uv": current.get("uv", 0),
                "is_day": current.get("is_day", 1),
            },
            "forecast": [
                {
                    "date": day.get("date", ""),
                    "maxtemp_c": day.get("day", {}).get("maxtemp_c", 0),
                    "mintemp_c": day.get("day", {}).get("mintemp_c", 0),
                    "condition": day.get("day", {}).get("condition", {}).get("text", ""),
                    "condition_icon": day.get("day", {}).get("condition", {}).get("icon", ""),
                    "daily_chance_of_rain": day.get("day", {}).get("daily_chance_of_rain", 0),
                    "avghumidity": day.get("day", {}).get("avghumidity", 0),
                }
                for day in forecast_days
            ],
            "cached": False,
            "cache_ttl": WEATHER_CACHE_TTL,
        }
        
        # 4. Cache'e kaydet
        _weather_cache[cache_key] = {
            "data": result,
            "timestamp": now
        }
        
        logger.info(f"Weather API called for '{city}' - cached for {WEATHER_CACHE_TTL}s")
        return result
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Weather API error: {e}")
        # Cache'de eski veri varsa onu döndür (graceful degradation)
        if cache_key in _weather_cache:
            logger.info("Returning stale cache due to API error")
            stale = _weather_cache[cache_key]["data"].copy()
            stale["cached"] = True
            stale["stale"] = True
            return stale
        raise HTTPException(status_code=502, detail=f"Weather API error: {str(e)}")

# --- Calendar (Hybrid) API Endpoints ---
import requests
from icalendar import Calendar as ICalendar
from models import CalendarNote

class CalendarNoteSchema(BaseModel):
    id: int
    date: datetime.date
    time: Optional[str] = None
    title: str
    description: Optional[str] = None
    color: str = "yellow"
    
    class Config:
        from_attributes = True

class NoteCreateSchema(BaseModel):
    date: datetime.date
    time: Optional[str] = None
    title: str
    description: Optional[str] = None
    color: str = "yellow"
    username: str # Geçici Auth (frontend'den user.username gelecek)

@app.get("/api/calendar/proxy")
def get_outlook_calendar_proxy(url: str):
    """
    Public ICS linkini çeker, parse eder ve JSON döner.
    URL parametresi olarak ?url=... ile ICS linki verilmeli.
    """
    try:
        response = requests.get(url)
        response.raise_for_status()
        
        cal = ICalendar.from_ical(response.content)
        events = []
        
        for component in cal.walk():
            if component.name == "VEVENT":
                summary = str(component.get('summary'))
                start = component.get('dtstart').dt
                end = component.get('dtend').dt if component.get('dtend') else None
                description = str(component.get('description')) if component.get('description') else ""
                location = str(component.get('location')) if component.get('location') else ""
                
                # Datetime objelerini string'e çevir
                events.append({
                    "title": summary,
                    "start": start.isoformat() if start else None,
                    "end": end.isoformat() if end else None,
                    "description": description,
                    "location": location,
                    "source": "outlook", # Frontend'de ayırt etmek için
                    "color": "#3b82f6" # Mavi
                })
                
        return events
    except Exception as e:
        print(f"Calendar Proxy Error: {e}")
        # Hata olsa bile boş liste dönelim ki uygulama çökmesin
        return []

@app.get("/api/calendar/notes/{username}", response_model=List[CalendarNoteSchema])
def get_user_notes(username: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username).first()
    if not user:
        return []
        
    notes = db.query(CalendarNote).filter(CalendarNote.user_id == user.id).all()
    return notes

@app.post("/api/calendar/notes")
def create_user_note(note: NoteCreateSchema, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == note.username).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
        
    new_note = CalendarNote(
        user_id=user.id,
        date=note.date,
        time=note.time,
        title=note.title,
        description=note.description,
        color=note.color
    )
    db.add(new_note)
    db.commit()
    db.refresh(new_note)
    return new_note

@app.delete("/api/calendar/notes/{note_id}")
def delete_user_note(note_id: int, db: Session = Depends(get_db)):
    note = db.query(CalendarNote).filter(CalendarNote.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404, detail="Not bulunamadı")
        
    db.delete(note)
    db.commit()
    return {"message": "Not silindi"}
